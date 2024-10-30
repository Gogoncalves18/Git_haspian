from dotenv import load_dotenv
from os import getenv
import os
from openpyxl import load_workbook


load_dotenv()

xls = getenv("table")
xls_path = getenv("loc_xls")
xls_imported = os.path.join(xls_path, xls)

# Carrega o arquivo excel
wb = load_workbook(xls_imported)
# Seleciona a primeira aba
ws = wb.worksheets[1]
ws_sspimport = wb['SSPImport']
ws_pcsops = wb['pcs_OPS']
ws_arranjos = wb['Arranjos']
# Leitura de linhas utilizadas e gravadas no metadados do arquivo, nao confiar
# neste parametro, necessario interar sobre ele
qtd_rows_ssp = ws_sspimport.max_row
qtd_rows_nest = ws_arranjos.max_row
qtd_rows_pcsops = ws_pcsops.max_row


# Parametro para contagem de linha
count_col = 1
count_row = 1
count_col_sspimport = 1
count_row_sspimport = 1
count_col_pcsops = 1
count_row_pcsops = 1
# Param para qtd de colunas + 1
qtd_real_col = 7
qtd_real_col_nest = 7
qtd_real_col_sspimport = 5
qtd_real_col_pcsops = 7


# Altera o valor da celula de ACAO
def atual_status(lin, ws):
    if ws == 'ssp':
        ws_sspimport.cell(row=lin, column=qtd_real_col_sspimport).value = 4
    elif ws == 'arranjos':
        ws_arranjos.cell(row=lin, column=qtd_real_col_nest).value = 4


def confirma_importacao(ops_to_excel):
    # Parametro para contagem de linha
    count_col_sspimport = 1
    count_rows_empty = 0
    global qtd_rows_ssp

    # Varre todas linhas (excluindo o cabecalho) da tabela ate o ultima linha
    # do metadado da tabela
    for lin in range(2, qtd_rows_ssp+1):
        if qtd_rows_ssp == 0:
            break
        # Intera sobre o numero de coluna indicado
        for col in range(1, qtd_real_col_sspimport):
            # Le a celula de cada linha e coluna
            val_cell = ws_sspimport.cell(
                row=lin, column=count_col_sspimport).value
            # Le apenas a primeira coluna e se esta preenchida
            if val_cell is not None and count_col_sspimport == 1:
                # Regra para encontrar a OF enviada para funcao em uma LIST
                # print(f'LISTA DENTRO DA CONFIRMACAO IMPORT {ops_to_excel}')
                if int(val_cell) in ops_to_excel:
                    print(f'>> Linha {lin} -- Col {col} '
                          f'-- ValCELL {val_cell}')
                    atual_status(lin, 'ssp')
            # Salva apos nao encontrar mais nada na primeira coluna
            if val_cell is None and count_col_sspimport == 1:
                count_rows_empty += 1
                wb.save(xls_imported)
                if count_rows_empty > 2:
                    qtd_rows_ssp = 0
                break
            # Intera a valor da coluna na mesma linha
            count_col_sspimport += 1
            # Controlador para validar se e a ultima coluna a ser
            if count_col_sspimport == qtd_real_col_sspimport:
                count_col_sspimport = 1


def confirma_importacao_nest(arranjos):
    # Parametro para contagem de linha
    count_col_sspimport = 1
    count_rows_empty = 0
    global qtd_rows_nest

    # Varre todas linhas (excluindo o cabecalho) da tabela ate o ultima linha
    # do metadado da tabela
    for lin in range(2, qtd_rows_nest+1):
        if qtd_rows_nest == 0:
            break
        # Intera sobre o numero de coluna indicado
        for col in range(1, qtd_real_col_nest):
            # Le a celula de cada linha e coluna
            val_cell = ws_arranjos.cell(
                row=lin, column=count_col_sspimport).value
            # Le apenas a primeira coluna e se esta preenchida
            if val_cell is not None and count_col_sspimport == 1:
                # Regra para encontrar a OF enviada para funcao em uma LIST
                # print(f'LISTA DENTRO DA CONFIRMACAO IMPORT {ops_to_excel}')
                if str(val_cell) in arranjos:
                    print(f'>> Linha {lin} -- Col {col} '
                          f'-- ValCELL {val_cell}')
                    atual_status(lin, 'arranjos')
            # Salva apos nao encontrar mais nada na primeira coluna
            if val_cell is None and count_col_sspimport == 1:
                count_rows_empty += 1
                wb.save(xls_imported)
                if count_rows_empty > 2:
                    qtd_rows_nest = 0
                break
            # Intera a valor da coluna na mesma linha
            count_col_sspimport += 1
            # Controlador para validar se e a ultima coluna a ser
            if count_col_sspimport == qtd_real_col_nest:
                count_col_sspimport = 1


if __name__ == '__main__':
    lt = [1000, 1000, 1001, 1001, 1001, 1002, 1002, 1002]
    lt_limpo = set(lt)
    confirma_importacao(lt_limpo)
